import multiprocessing

import pandas as pd
import pandas_ta as ta
from backtesting import Backtest
from backtesting import Strategy

from DataConverter import CommonUtils
from DataGetter import getNifty500StocksMetaData, allStockMetaDataWithData
from openpyxl import Workbook, load_workbook


class RsiOscillator(Strategy):
    buyRSIvalue = 20
    buyRSIvalueDelta = 2.5
    rsi_window = 14
    tp = 40
    sl = 20

    def init(self):
        self.rsi = self.I(ta.rsi, pd.Series(self.data.df['Close']), self.rsi_window)

    def next(self):
        price = self.data['Close'][-1]

        if CommonUtils.betweenDeltaInclusive(self.rsi[-1], self.buyRSIvalue, self.buyRSIvalueDelta,
                                             self.buyRSIvalueDelta):
            if not self.position.is_short:
                self.sell(sl=price * (1 + (self.sl / 1000)), tp=price * (1 - (self.tp / 1000)))


availableKeys = ["Return [%]", "_strategy", "Start", "End", "Duration", "Exposure Time [%]", "Equity Final [$]",
                 "Equity Peak [$]", "Buy & Hold Return [%]",
                 "Return (Ann.) [%]", "Volatility (Ann.) [%]", "Sharpe Ratio", "Sortino Ratio", "Calmar Ratio",
                 "Max. Drawdown [%]", "Avg. Drawdown [%]",
                 "Max. Drawdown Duration", "Avg. Drawdown Duration", "# Trades", "Win Rate [%]", "Best Trade [%]",
                 "Best Trade [%]",
                 "Avg. Trade [%]", "Avg. Trade [%]", "Avg. Trade [%]", "Profit Factor", "Expectancy [%]", "SQN"]


def runEachCase(row):
    try:
        multiprocessing.set_start_method('fork')
    except Exception as e:
        print("skip")

    bt = Backtest(row[-1], RsiOscillator, cash=10_000)
    stats = bt.optimize(
        buyRSIvalue=range(20, 85, 1),
        rsi_window=range(10, 20, 1),
        maximize='Return [%]')
    # bt.plot()
    stockReturn = ((row[-1]['Close'].iloc[-1] - row[-1]['Close'].iloc[0]) / row[-1]['Close'].iloc[0]) * 100
    data = [row[0], row[2], stockReturn]
    for key in availableKeys:
        val = stats[key]
        try:
            val = float(stats[key])
        except Exception as e:
            val = str(stats[key])
        data.append(val)
    print(data)
    addToExcel(data, header=False)


def runAllCases(nifty500StocksMetaDataWithData):
    with multiprocessing.Pool() as pool:
        pool.map(runEachCase, nifty500StocksMetaDataWithData)


file_path = "example.xlsx"


def addToExcel(row, header=False):
    if header:
        workbook = Workbook()
        sheet = workbook.active
        headers = ["stockCode", "stockSymbol", "stockReturn"]
        for key in availableKeys:
            headers.append(key)
        sheet.append(headers)

    else:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        sheet.append(row)

    try:
        workbook.save(file_path)
        print("Row added successfully to:", file_path)
    except Exception as e:
        print(f"An error occurred while saving the file: {e}")


if __name__ == "__main__":
    nifty500StocksMetaData = getNifty500StocksMetaData()
    start_date = '2023-10-08 09:16'
    end_date = '2024-05-11 09:16'
    timeFrameTickSize = "5m"  # 1m , 5m , 15m, 30m, 1d
    file_path = "example.xlsx"
    addToExcel([], header=True)
    nifty500StocksMetaDataWithData = allStockMetaDataWithData(nifty500StocksMetaData, timeFrameTickSize, start_date,
                                                              end_date)
    runAllCases(nifty500StocksMetaDataWithData)
